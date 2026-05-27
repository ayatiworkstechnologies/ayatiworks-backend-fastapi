"""
Employee API routes.
Employee CRUD and management.
"""


from fastapi import APIRouter, Depends, HTTPException, Query, status
from sqlalchemy.orm import Session

from app.api.deps import PermissionChecker, get_current_active_user, get_user_permissions
from app.core.exceptions import PermissionDeniedError, ResourceNotFoundError
from app.core.permissions import check_permission
from app.database import get_db
from app.models.auth import User
from app.schemas.common import MessageResponse, PaginatedResponse
from app.schemas.employee import (
    EmployeeCreate,
    EmployeeDocumentResponse,
    EmployeeListResponse,
    EmployeeResponse,
    EmployeeTeamResponse,
    EmployeeUpdate,
)
from app.services.employee_service import EmployeeService

router = APIRouter(prefix="/employees", tags=["Employees"])


def build_employee_response(employee) -> EmployeeResponse:
    """Helper function to build EmployeeResponse from Employee model."""

    # Build teams list
    teams_data = []
    if hasattr(employee, 'team_memberships') and employee.team_memberships:
        for tm in employee.team_memberships:
            if tm.team:
                teams_data.append(EmployeeTeamResponse(
                    id=tm.team.id,
                    name=tm.team.name,
                    code=tm.team.code,
                    team_type=tm.team.team_type,
                    role=tm.role,
                    joined_date=tm.joined_date
                ))

    return EmployeeResponse(
        id=employee.id,
        user_id=employee.user_id,
        employee_code=employee.employee_code,
        company_id=employee.company_id,
        branch_id=employee.branch_id,
        department_id=employee.department_id,
        designation_id=employee.designation_id,
        manager_id=employee.manager_id,
        joining_date=employee.joining_date,
        probation_end_date=employee.probation_end_date,
        confirmation_date=employee.confirmation_date,
        employment_type=employee.employment_type,
        employment_status=employee.employment_status,
        work_mode=employee.work_mode,
        shift_id=employee.shift_id,
        is_active=employee.is_active,
        date_of_birth=employee.date_of_birth,
        gender=employee.gender,
        blood_group=employee.blood_group,
        marital_status=employee.marital_status,
        nationality=employee.nationality,
        personal_email=employee.personal_email,
        personal_phone=employee.personal_phone,
        emergency_contact_name=employee.emergency_contact_name,
        emergency_contact_phone=employee.emergency_contact_phone,
        emergency_contact_relation=employee.emergency_contact_relation,
        current_address=employee.current_address,
        permanent_address=employee.permanent_address,
        city=employee.city,
        state=employee.state,
        country=employee.country,
        postal_code=employee.postal_code,
        created_at=employee.created_at,
        updated_at=employee.updated_at,
        user={
            "id": employee.user.id,
            "email": employee.user.email,
            "first_name": employee.user.first_name,
            "last_name": employee.user.last_name,
            "phone": employee.user.phone,
            "avatar": employee.user.avatar
        } if employee.user else None,
        department_name=employee.department.name if employee.department else None,
        designation_name=employee.designation.name if employee.designation else None,
        manager_name=employee.manager.user.full_name if employee.manager and employee.manager.user else None,
        teams=teams_data
    )


@router.get("", response_model=PaginatedResponse[EmployeeListResponse])
async def list_employees(
    company_id: int | None = None,
    branch_id: int | None = None,
    department_id: int | None = None,
    designation_id: int | None = None,
    status: str | None = None,
    search: str | None = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """List all employees with filters and pagination."""
    service = EmployeeService(db)
    permissions = get_user_permissions(current_user, db)

    if not check_permission(permissions, "employee.view_all"):
        if not check_permission(permissions, "employee.view"):
            raise PermissionDeniedError("Permission denied: employee.view")

        employee = service.get_by_user_id(current_user.id)
        if not employee:
            return PaginatedResponse.create([], 0, page, page_size)

        matches_search = not search or any(
            term and search.lower() in term.lower()
            for term in [
                employee.employee_code,
                employee.user.first_name if employee.user else "",
                employee.user.last_name if employee.user else "",
                employee.user.email if employee.user else "",
            ]
        )
        matches_status = not status or employee.employment_status == status
        matches_company = not company_id or employee.company_id == company_id
        matches_branch = not branch_id or employee.branch_id == branch_id
        matches_department = not department_id or employee.department_id == department_id
        matches_designation = not designation_id or employee.designation_id == designation_id

        if not all([
            matches_search,
            matches_status,
            matches_company,
            matches_branch,
            matches_department,
            matches_designation,
        ]):
            return PaginatedResponse.create([], 0, page, page_size)

        item = EmployeeListResponse(
            id=employee.id,
            user_id=employee.user_id,
            employee_code=employee.employee_code,
            first_name=employee.user.first_name if employee.user else "",
            last_name=employee.user.last_name if employee.user else None,
            email=employee.user.email if employee.user else "",
            avatar=employee.user.avatar if employee.user else None,
            department_name=employee.department.name if employee.department else None,
            designation_name=employee.designation.name if employee.designation else None,
            employment_status=employee.employment_status,
            is_active=employee.is_active,
        )
        return PaginatedResponse.create([item], 1, 1, page_size)

    employees, total = service.get_all(
        company_id=company_id,
        branch_id=branch_id,
        department_id=department_id,
        designation_id=designation_id,
        status=status,
        search=search,
        page=page,
        page_size=page_size
    )

    # Convert to response format
    items = []
    for emp in employees:
        items.append(EmployeeListResponse(
            id=emp.id,
            user_id=emp.user_id,
            employee_code=emp.employee_code,
            first_name=emp.user.first_name if emp.user else "",
            last_name=emp.user.last_name if emp.user else None,
            email=emp.user.email if emp.user else "",
            avatar=emp.user.avatar if emp.user else None,
            department_name=emp.department.name if emp.department else None,
            designation_name=emp.designation.name if emp.designation else None,
            employment_status=emp.employment_status,
            is_active=emp.is_active
        ))

    return PaginatedResponse.create(items, total, page, page_size)


@router.get("/me", response_model=EmployeeResponse)
async def get_my_profile(
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get current user's employee profile."""
    service = EmployeeService(db)

    employee = service.get_by_user_id(current_user.id)

    if not employee:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Employee profile not found"
        )

    return build_employee_response(employee)


@router.get("/next-code", response_model=dict)
async def get_next_employee_code(
    prefix: str | None = None,
    current_user: User = Depends(PermissionChecker("employee.create")),
    db: Session = Depends(get_db)
):
    """
    Get the next available employee code.
    Useful for auto-generating IDs in the frontend.
    """
    service = EmployeeService(db)
    code = service.generate_employee_code(prefix=prefix)
    return {"code": code}


@router.get("/code/{code}", response_model=EmployeeResponse)
async def get_employee_by_code(
    code: str,
    current_user: User = Depends(PermissionChecker("employee.view")),
    db: Session = Depends(get_db)
):
    """Get employee by employee code (e.g., AW0001)."""
    service = EmployeeService(db)

    employee = service.get_by_code(code.upper())

    if not employee:
        raise ResourceNotFoundError("Employee", code)

    return build_employee_response(employee)


@router.get("/{employee_id}", response_model=EmployeeResponse)
async def get_employee(
    employee_id: int,
    current_user: User = Depends(PermissionChecker("employee.view")),
    db: Session = Depends(get_db)
):
    """Get employee by ID."""
    service = EmployeeService(db)

    employee = service.get_by_id(employee_id)

    if not employee:
        raise ResourceNotFoundError("Employee", employee_id)

    return build_employee_response(employee)


@router.post("", response_model=EmployeeResponse, status_code=status.HTTP_201_CREATED)
async def create_employee(
    data: EmployeeCreate,
    current_user: User = Depends(PermissionChecker("employee.create")),
    db: Session = Depends(get_db)
):
    """
    Create a new employee.
    If user_id is not provided, a new user account will be created.
    Employee code (like AW0001) is auto-generated.
    A welcome email will be sent to the new employee.
    """
    from app.services.email_service import email_service, employee_welcome_email

    service = EmployeeService(db)

    # Store password before hashing (for email)
    raw_password = data.password

    try:
        employee = service.create(data, created_by=current_user.id)
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )

    # Send welcome email
    try:
        department_name = employee.department.name if employee.department else "N/A"
        designation_name = employee.designation.name if employee.designation else "N/A"

        subject, html_content = employee_welcome_email(
            first_name=employee.user.first_name,
            last_name=employee.user.last_name or "",
            email=employee.user.email,
            employee_code=employee.employee_code,
            department=department_name,
            designation=designation_name,
            joining_date=str(employee.joining_date),
            password=raw_password  # Only include if new user was created
        )

        from app.config import settings

        if settings.REDIS_URL:
            from app.tasks.email_tasks import send_email_async

            send_email_async.delay(
                to_email=employee.user.email,
                subject=subject,
                html_content=html_content
            )
        else:
            email_service.send_email(
                to_email=employee.user.email,
                subject=subject,
                html_content=html_content
            )
    except Exception as e:
        # Log email error but don't fail the request
        import logging
        logging.error(f"Failed to send welcome email: {e}")

    return build_employee_response(employee)


@router.put("/{employee_id}", response_model=EmployeeResponse)
async def update_employee(
    employee_id: int,
    data: EmployeeUpdate,
    current_user: User = Depends(PermissionChecker("employee.edit")),
    db: Session = Depends(get_db)
):
    """Update an employee."""
    service = EmployeeService(db)

    employee = service.update(employee_id, data, updated_by=current_user.id)

    if not employee:
        raise ResourceNotFoundError("Employee", employee_id)

    return build_employee_response(employee)


@router.delete("/{employee_id}", response_model=MessageResponse)
async def delete_employee(
    employee_id: int,
    current_user: User = Depends(PermissionChecker("employee.delete")),
    db: Session = Depends(get_db)
):
    """Delete an employee (soft delete)."""
    service = EmployeeService(db)

    if not service.delete(employee_id, deleted_by=current_user.id):
        raise ResourceNotFoundError("Employee", employee_id)

    return MessageResponse(message="Employee deleted successfully")


@router.get("/{employee_id}/team", response_model=list[EmployeeListResponse])
async def get_team_members(
    employee_id: int,
    current_user: User = Depends(PermissionChecker("employee.view")),
    db: Session = Depends(get_db)
):
    """Get all team members under an employee (manager)."""
    service = EmployeeService(db)

    employees = service.get_team_members(employee_id)

    items = []
    for emp in employees:
        items.append(EmployeeListResponse(
            id=emp.id,
            user_id=emp.user_id,
            employee_code=emp.employee_code,
            first_name=emp.user.first_name if emp.user else "",
            last_name=emp.user.last_name if emp.user else None,
            email=emp.user.email if emp.user else "",
            department_name=emp.department.name if emp.department else None,
            designation_name=emp.designation.name if emp.designation else None,
            employment_status=emp.employment_status,
            is_active=emp.is_active
        ))
    return items


# Document endpoints
@router.get("/{employee_id}/documents", response_model=list[EmployeeDocumentResponse])
async def get_employee_documents(
    employee_id: int,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """Get all documents for an employee."""
    service = EmployeeService(db)

    # Check if user can view this employee's documents
    employee = service.get_by_id(employee_id)
    if not employee:
        raise ResourceNotFoundError("Employee", employee_id)

    documents = service.get_documents(employee_id)

    return [EmployeeDocumentResponse.model_validate(doc) for doc in documents]


@router.post("/{employee_id}/documents/{document_id}/verify", response_model=EmployeeDocumentResponse)
async def verify_document(
    employee_id: int,
    document_id: int,
    current_user: User = Depends(PermissionChecker("employee.edit")),
    db: Session = Depends(get_db)
):
    """Mark a document as verified."""
    service = EmployeeService(db)

    document = service.verify_document(document_id, verified_by=current_user.id)

    if not document:
        raise ResourceNotFoundError("Document", document_id)

    return EmployeeDocumentResponse.model_validate(document)


@router.get("/export/{format}")
async def export_employees(
    format: str,
    department_id: int | None = None,
    status: str | None = None,
    current_user: User = Depends(PermissionChecker("employee.view_all")),
    db: Session = Depends(get_db)
):
    """
    Export employees to CSV, Excel, or PDF.
    Requires employee.view_all permission.
    """
    import csv
    from io import BytesIO, StringIO

    from fastapi.responses import StreamingResponse

    if format not in ("csv", "excel", "pdf"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Format must be csv, excel, or pdf"
        )

    service = EmployeeService(db)

    employees, _ = service.get_all(
        department_id=department_id,
        status=status,
        page=1,
        page_size=10000  # Max export limit
    )

    headers = [
        "Employee Code", "First Name", "Last Name", "Email",
        "Department", "Designation", "Joining Date",
        "Employment Type", "Employment Status", "Work Mode", "Phone"
    ]

    rows = []
    for emp in employees:
        rows.append([
            emp.employee_code,
            emp.user.first_name if emp.user else "",
            emp.user.last_name if emp.user else "",
            emp.user.email if emp.user else "",
            emp.department.name if emp.department else "",
            emp.designation.name if emp.designation else "",
            str(emp.joining_date) if emp.joining_date else "",
            emp.employment_type or "",
            emp.employment_status or "",
            emp.work_mode or "",
            emp.personal_phone or ""
        ])

    if format == "csv":
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=employees.csv"}
        )

    if format == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for r, row in enumerate(rows, 2):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=employees.xlsx"}
        )

    # PDF
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    table_data = [headers] + rows
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    doc.build([table])
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=employees.pdf"}
    )


@router.post("/bulk-delete")
async def bulk_delete_employees(
    employee_ids: list[int],
    current_user: User = Depends(PermissionChecker("employee.delete")),
    db: Session = Depends(get_db)
):
    """
    Bulk delete multiple employees.
    Requires employee.delete permission.
    Optimized: single batch update instead of sequential loop.
    """
    if not employee_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No employee IDs provided"
        )

    from datetime import datetime
    from app.models.employee import Employee

    deleted_count = db.query(Employee).filter(
        Employee.id.in_(employee_ids),
        Employee.is_deleted == False
    ).update(
        {
            Employee.is_deleted: True,
            Employee.deleted_at: datetime.utcnow(),
            Employee.deleted_by: current_user.id
        },
        synchronize_session="fetch"
    )
    db.commit()

    return {"message": f"{deleted_count} employees deleted successfully", "count": deleted_count}
