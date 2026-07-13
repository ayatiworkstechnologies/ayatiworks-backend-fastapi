"""
Client Modules API routes.
SMTP config, mail templates, dynamic modules, and module records.
Admin-only access — clients cannot login.

Optimised:
- aiosmtplib for non-blocking async SMTP
- Helper builders to eliminate repeated serialisation code
- Batch SQL queries (no N+1)
- search support on list_module_records
- input sanitisation on public endpoints
- all imports moved to top-level
"""

import csv
import io
import json
import logging
import re
import secrets
import ssl
from datetime import UTC, datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import aiosmtplib
from fastapi import APIRouter, Depends, Header, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from jinja2 import BaseLoader, Environment
from sqlalchemy import func, inspect, text
from sqlalchemy.orm import Session

from app.api.deps import PermissionChecker
from app.database import get_db
from app.models.auth import User
from app.models.client import Client
from app.models.client_module import (
    ClientMailTemplate,
    ClientModule,
    ClientModuleRecord,
    ClientSmtpConfig,
)
from app.schemas.client_module import (
    ClientMailTemplateCreate,
    ClientMailTemplateListResponse,
    ClientMailTemplateResponse,
    ClientMailTemplateUpdate,
    ClientModuleCreate,
    ClientModuleListResponse,
    ClientModuleRecordCreate,
    ClientModuleRecordResponse,
    ClientModuleRecordUpdate,
    ClientModuleResponse,
    ClientModuleUpdate,
    ClientSendEmailRequest,
    ClientSmtpConfigCreate,
    ClientSmtpConfigResponse,
)
from app.schemas.common import MessageResponse, PaginatedResponse
from app.services.email_service import email_service

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Client Modules"])


# ============================================================
# Shared Helpers
# ============================================================

def _get_client_or_404(client_id: int, db: Session) -> Client:
    """Get client by ID or raise 404."""
    client = db.query(Client).filter(
        Client.id == client_id,
        Client.is_deleted == False,
    ).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    return client


def _slugify(text: str) -> str:
    """Convert text to URL-friendly slug."""
    text = text.lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_]+', '_', text)
    return text


def _build_smtp_response(config: ClientSmtpConfig) -> ClientSmtpConfigResponse:
    """Build ClientSmtpConfigResponse from model — avoids repeated field mapping."""
    return ClientSmtpConfigResponse(
        id=config.id,
        client_id=config.client_id,
        host=config.host,
        port=config.port,
        username=config.username,
        password_set=bool(config.password),
        from_email=config.from_email,
        from_name=config.from_name,
        use_tls=config.use_tls,
        created_at=config.created_at,
        updated_at=config.updated_at,
    )


def _build_template_response(template: ClientMailTemplate) -> ClientMailTemplateResponse:
    """Build ClientMailTemplateResponse from model."""
    return ClientMailTemplateResponse(
        id=template.id,
        client_id=template.client_id,
        name=template.name,
        subject=template.subject,
        html_body=template.html_body,
        to_email=template.to_email,
        from_email=template.from_email,
        cc_email=template.cc_email,
        bcc_email=template.bcc_email,
        variables=template.variables,
        created_at=template.created_at,
        updated_at=template.updated_at,
    )


def _build_module_response(module: ClientModule, record_count: int = 0) -> ClientModuleResponse:
    """Build ClientModuleResponse from model."""
    return ClientModuleResponse(
        id=module.id,
        client_id=module.client_id,
        name=module.name,
        slug=module.slug,
        description=module.description,
        icon=module.icon,
        fields=module.fields or [],
        mail_template_id=getattr(module, "mail_template_id", None),
        record_count=record_count,
        created_at=module.created_at,
        updated_at=module.updated_at,
    )


def _build_record_response(record: ClientModuleRecord, email_sent: bool = False) -> ClientModuleRecordResponse:
    """Build ClientModuleRecordResponse from model."""
    return ClientModuleRecordResponse(
        id=record.id,
        module_id=record.module_id,
        data=record.data or {},
        email_sent=email_sent,
        created_at=record.created_at,
        updated_at=record.updated_at,
    )


def _parse_record_data(value) -> dict:
    """Normalize JSON data returned by different DB drivers."""
    if isinstance(value, dict):
        return value
    if isinstance(value, bytes | bytearray):
        value = value.decode("utf-8")
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
            return parsed if isinstance(parsed, dict) else {}
        except json.JSONDecodeError:
            return {}
    return {}


def _get_table_columns(db: Session, table_name: str) -> set[str]:
    """Return the current DB columns for a table."""
    return {column["name"] for column in inspect(db.get_bind()).get_columns(table_name)}


def _legacy_safe_record_response(row) -> ClientModuleRecordResponse:
    """Build a record response from a raw row that may not contain every model column."""
    mapping = row._mapping
    return ClientModuleRecordResponse(
        id=mapping["id"],
        module_id=mapping["module_id"],
        data=_parse_record_data(mapping.get("data")),
        email_sent=False,
        created_at=mapping.get("created_at"),
        updated_at=mapping.get("updated_at"),
    )


def _get_record_count(module_id: int, db: Session) -> int:
    """Fetch record count for one module efficiently."""
    return db.query(func.count(ClientModuleRecord.id)).filter(
        ClientModuleRecord.module_id == module_id,
        ClientModuleRecord.is_deleted == False,
    ).scalar() or 0


def _validate_required_fields(module: ClientModule, data: dict):
    """Validate submitted data against module field definitions."""
    for field_def in (module.fields or []):
        if field_def.get("required") and field_def.get("name") not in data:
            label = field_def.get("label", field_def.get("name"))
            raise HTTPException(status_code=400, detail=f"Field '{label}' is required")


# ============================================================
# SMTP Config Endpoints
# ============================================================

@router.get("/clients/{client_id}/smtp", response_model=ClientSmtpConfigResponse | None)
async def get_smtp_config(
    client_id: int,
    current_user: User = Depends(PermissionChecker("mail.view")),
    db: Session = Depends(get_db),
):
    """Get SMTP configuration for a client."""
    _get_client_or_404(client_id, db)
    config = db.query(ClientSmtpConfig).filter(
        ClientSmtpConfig.client_id == client_id,
        ClientSmtpConfig.is_deleted == False,
    ).first()
    return _build_smtp_response(config) if config else None


@router.post("/clients/{client_id}/smtp", response_model=ClientSmtpConfigResponse, status_code=status.HTTP_201_CREATED)
async def create_or_update_smtp_config(
    client_id: int,
    data: ClientSmtpConfigCreate,
    current_user: User = Depends(PermissionChecker("mail.manage")),
    db: Session = Depends(get_db),
):
    """Create or update SMTP configuration for a client."""
    _get_client_or_404(client_id, db)

    config = db.query(ClientSmtpConfig).filter(
        ClientSmtpConfig.client_id == client_id,
        ClientSmtpConfig.is_deleted == False,
    ).first()

    if config:
        for field, value in data.model_dump(exclude_unset=True).items():
            setattr(config, field, value)
        config.updated_by = current_user.id
    else:
        config = ClientSmtpConfig(
            client_id=client_id,
            **data.model_dump(),
            created_by=current_user.id,
        )
        db.add(config)

    db.commit()
    db.refresh(config)
    return _build_smtp_response(config)


@router.delete("/clients/{client_id}/smtp", response_model=MessageResponse)
async def delete_smtp_config(
    client_id: int,
    current_user: User = Depends(PermissionChecker("mail.manage")),
    db: Session = Depends(get_db),
):
    """Delete a client's SMTP configuration (revert to system default)."""
    _get_client_or_404(client_id, db)
    config = db.query(ClientSmtpConfig).filter(
        ClientSmtpConfig.client_id == client_id,
        ClientSmtpConfig.is_deleted == False,
    ).first()
    if config:
        config.soft_delete(current_user.id)
        db.commit()
    return MessageResponse(message="SMTP configuration removed. Reverted to System SMTP.")


@router.post("/clients/{client_id}/smtp/test", response_model=MessageResponse)
async def test_smtp_config(
    client_id: int,
    data: ClientSmtpConfigCreate,
    current_user: User = Depends(PermissionChecker("mail.manage")),
    db: Session = Depends(get_db),
):
    """Test SMTP configuration — async non-blocking connection test."""
    _get_client_or_404(client_id, db)

    try:
        context = ssl.create_default_context()
        if data.port == 465:
            smtp = aiosmtplib.SMTP(
                hostname=data.host,
                port=data.port,
                use_tls=True,
                tls_context=context,
                timeout=10,
            )
        else:
            smtp = aiosmtplib.SMTP(
                hostname=data.host,
                port=data.port,
                timeout=10,
            )

        await smtp.connect()
        if data.port != 465 and data.use_tls:
            await smtp.starttls(tls_context=context)
        await smtp.login(data.username, data.password)
        await smtp.quit()

        return MessageResponse(message="SMTP connection successful!")

    except aiosmtplib.SMTPAuthenticationError:
        raise HTTPException(status_code=400, detail="Authentication failed. Check username and password.")
    except (TimeoutError, aiosmtplib.SMTPConnectError, OSError):
        raise HTTPException(status_code=400, detail="Could not connect to the server. Check host and port.")
    except Exception as e:
        logger.error(f"SMTP Test Error for client {client_id}: {e}")
        raise HTTPException(status_code=400, detail=f"Connection failed: {str(e)}")


# ============================================================
# Mail Template Endpoints
# ============================================================

@router.get("/clients/{client_id}/mail-templates", response_model=PaginatedResponse[ClientMailTemplateListResponse])
async def list_mail_templates(
    client_id: int,
    search: str | None = Query(None, description="Filter by name or subject"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(PermissionChecker("mail.view")),
    db: Session = Depends(get_db),
):
    """List mail templates for a client."""
    _get_client_or_404(client_id, db)

    query = db.query(ClientMailTemplate).filter(
        ClientMailTemplate.client_id == client_id,
        ClientMailTemplate.is_deleted == False,
    )
    if search:
        like = f"%{search}%"
        query = query.filter(
            ClientMailTemplate.name.ilike(like) | ClientMailTemplate.subject.ilike(like)
        )

    total = query.count()
    templates = query.order_by(ClientMailTemplate.created_at.desc()).offset(
        (page - 1) * page_size
    ).limit(page_size).all()

    items = [
        ClientMailTemplateListResponse(
            id=t.id,
            client_id=t.client_id,
            name=t.name,
            subject=t.subject,
            variables=t.variables,
        )
        for t in templates
    ]
    return PaginatedResponse.create(items, total, page, page_size)


@router.get("/clients/{client_id}/mail-templates/{template_id}", response_model=ClientMailTemplateResponse)
async def get_mail_template(
    client_id: int,
    template_id: int,
    current_user: User = Depends(PermissionChecker("mail.view")),
    db: Session = Depends(get_db),
):
    """Get a specific mail template."""
    _get_client_or_404(client_id, db)
    template = db.query(ClientMailTemplate).filter(
        ClientMailTemplate.id == template_id,
        ClientMailTemplate.client_id == client_id,
        ClientMailTemplate.is_deleted == False,
    ).first()
    if not template:
        raise HTTPException(status_code=404, detail="Mail template not found")
    return _build_template_response(template)


@router.post("/clients/{client_id}/mail-templates", response_model=ClientMailTemplateResponse, status_code=status.HTTP_201_CREATED)
async def create_mail_template(
    client_id: int,
    data: ClientMailTemplateCreate,
    current_user: User = Depends(PermissionChecker("mail.manage")),
    db: Session = Depends(get_db),
):
    """Create a new mail template."""
    _get_client_or_404(client_id, db)
    template = ClientMailTemplate(
        client_id=client_id,
        **data.model_dump(),
        created_by=current_user.id,
    )
    db.add(template)
    db.commit()
    db.refresh(template)
    return _build_template_response(template)


@router.put("/clients/{client_id}/mail-templates/{template_id}", response_model=ClientMailTemplateResponse)
async def update_mail_template(
    client_id: int,
    template_id: int,
    data: ClientMailTemplateUpdate,
    current_user: User = Depends(PermissionChecker("mail.manage")),
    db: Session = Depends(get_db),
):
    """Update a mail template."""
    _get_client_or_404(client_id, db)
    template = db.query(ClientMailTemplate).filter(
        ClientMailTemplate.id == template_id,
        ClientMailTemplate.client_id == client_id,
        ClientMailTemplate.is_deleted == False,
    ).first()
    if not template:
        raise HTTPException(status_code=404, detail="Mail template not found")

    for field, value in data.model_dump(exclude_unset=True).items():
        setattr(template, field, value)
    template.updated_by = current_user.id
    db.commit()
    db.refresh(template)
    return _build_template_response(template)


@router.delete("/clients/{client_id}/mail-templates/{template_id}", response_model=MessageResponse)
async def delete_mail_template(
    client_id: int,
    template_id: int,
    current_user: User = Depends(PermissionChecker("mail.manage")),
    db: Session = Depends(get_db),
):
    """Delete a mail template."""
    _get_client_or_404(client_id, db)
    template = db.query(ClientMailTemplate).filter(
        ClientMailTemplate.id == template_id,
        ClientMailTemplate.client_id == client_id,
        ClientMailTemplate.is_deleted == False,
    ).first()
    if not template:
        raise HTTPException(status_code=404, detail="Mail template not found")

    template.soft_delete(current_user.id)
    db.commit()
    return MessageResponse(message="Mail template deleted successfully")


# ============================================================
# Send Email Endpoint
# ============================================================

def _render_jinja(template_str: str, variables: dict) -> str:
    """Render a Jinja2 template string with given variables."""
    jinja_env = Environment(loader=BaseLoader(), autoescape=True)
    try:
        return jinja_env.from_string(template_str).render(**variables)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Template rendering error: {str(e)}")


@router.post("/clients/{client_id}/send-email", response_model=MessageResponse)
async def send_client_email(
    client_id: int,
    data: ClientSendEmailRequest,
    current_user: User = Depends(PermissionChecker("mail.send")),
    db: Session = Depends(get_db),
):
    """Send email using client's SMTP config and optional template."""
    _get_client_or_404(client_id, db)

    smtp_config = db.query(ClientSmtpConfig).filter(
        ClientSmtpConfig.client_id == client_id,
        ClientSmtpConfig.is_deleted == False,
    ).first()

    subject = data.subject
    html_body = data.html_body

    if data.template_id:
        tmpl = db.query(ClientMailTemplate).filter(
            ClientMailTemplate.id == data.template_id,
            ClientMailTemplate.client_id == client_id,
            ClientMailTemplate.is_deleted == False,
        ).first()
        if not tmpl:
            raise HTTPException(status_code=404, detail="Mail template not found")
        subject = data.subject or tmpl.subject
        html_body = data.html_body or tmpl.html_body

    if not subject or not html_body:
        raise HTTPException(status_code=400, detail="Subject and body are required")

    if data.variables:
        subject = _render_jinja(subject, data.variables)
        html_body = _render_jinja(html_body, data.variables)

    await _send_email_logic(client_id, data.to_email, subject, html_body, db, smtp_config, data.cc, data.bcc)
    return MessageResponse(message=f"Email sent successfully to {data.to_email}")


async def _send_email_logic(
    client_id: int,
    to_email: str,
    subject: str,
    html_body: str,
    db: Session,
    smtp_config: ClientSmtpConfig | None = None,
    cc: list[str] | None = None,
    bcc: list[str] | None = None,
):
    """Internal helper to send email via Client SMTP or System Fallback."""
    if not smtp_config:
        # Try fetching if not provided
        smtp_config = db.query(ClientSmtpConfig).filter(
            ClientSmtpConfig.client_id == client_id,
            ClientSmtpConfig.is_deleted == False,
        ).first()

    # ---- Use System SMTP if no client config ----
    if not smtp_config:
        try:
            success = email_service.send_email(
                to_email=to_email,
                subject=subject,
                html_content=html_body,
                cc=cc,
                bcc=bcc,
            )
            if not success:
                raise Exception("System email service returned failure")
            return
        except Exception as e:
            logger.error(f"System email sending failed for client {client_id}: {e}")
            raise HTTPException(status_code=500, detail=f"Failed to send email via System SMTP: {str(e)}")

    # ---- Use Client Custom SMTP (aiosmtplib - non-blocking) ----
    try:
        # Optionally wrap in system base template
        try:
            wrapped_body = email_service.render_template(
                "email/client_custom.html",
                {"custom_content": html_body, "title": subject},
            )
        except Exception:
            wrapped_body = html_body

        from_display = (
            f"{smtp_config.from_name} <{smtp_config.from_email}>"
            if smtp_config.from_name
            else smtp_config.from_email
        )

        recipients = [to_email]
        if cc:
            recipients.extend(cc)
        if bcc:
            recipients.extend(bcc)

        context = ssl.create_default_context()
        use_tls_on_connect = smtp_config.port == 465

        # Create message first to fail fast on encoding errors
        message = MIMEMultipart("alternative")
        message["Subject"] = subject
        message["From"] = from_display
        message["To"] = to_email
        if cc:
            message["Cc"] = ", ".join(cc)
        message.attach(MIMEText(wrapped_body, "html"))

        smtp = aiosmtplib.SMTP(
            hostname=smtp_config.host,
            port=smtp_config.port,
            use_tls=use_tls_on_connect,
            tls_context=context,
            timeout=15,
        )
        await smtp.connect()
        if not use_tls_on_connect and smtp_config.use_tls:
            await smtp.starttls(tls_context=context)
        await smtp.login(smtp_config.username, smtp_config.password)

        await smtp.send_message(message, recipients=recipients)
        await smtp.quit()

        logger.info(f"Client email sent to {to_email} from client {client_id}")

    except aiosmtplib.SMTPAuthenticationError:
        raise HTTPException(status_code=400, detail="SMTP authentication failed. Verify SMTP credentials.")
    except aiosmtplib.SMTPException as e:
        logger.error(f"SMTP error for client {client_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to send email: {str(e)}")
    except Exception as e:
        logger.error(f"Email sending failed for client {client_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to send email: {str(e)}")


async def _trigger_module_email(client_id: int, module: ClientModule, record_data: dict, db: Session) -> bool:
    """Trigger automated email if module has a template. Returns True if email sent."""
    if not module.mail_template_id:
        return False

    template = db.query(ClientMailTemplate).filter(
        ClientMailTemplate.id == module.mail_template_id,
        ClientMailTemplate.is_deleted == False,
    ).first()

    if not template:
        logger.warning(f"Module {module.id} references missing template {module.mail_template_id}")
        return False

    # Determine recipient: Prioritize Static To Email in template
    recipient_email = template.to_email
    if recipient_email:
        try:
            recipient_email = _render_jinja(recipient_email, record_data)
        except Exception:
            logger.warning(f"Failed to render recipient email template: {recipient_email}")
            pass

    if not recipient_email:
        # Fallback: check record data
        for key, value in record_data.items():
            if key.lower() in ['email', 'to_email', 'contact_email', 'mail']:
                recipient_email = value
                break

    if not recipient_email:
        # Fallback: check schema for field of type 'email'
        for field in (module.fields or []):
            if field.get("type") == "email" and field.get("name") in record_data:
                recipient_email = record_data[field["name"]]
                break

    if not recipient_email:
        logger.warning(f"Module {module.id} template {template.id} has no To Email and no email found in record")
        return False

    # Render template with record data
    try:
        subject = _render_jinja(template.subject, record_data)
        html_body = _render_jinja(template.html_body, record_data)

        # Auto-render file/image fields as HTML in email body
        attachments_html = _build_file_image_html(module.fields or [], record_data)
        if attachments_html:
            html_body += attachments_html

        # Send non-blocking (fire and forget from client perspective, but we await here for reliability)
        await _send_email_logic(client_id, recipient_email, subject, html_body, db)
        return True
    except Exception as e:
        logger.error(f"Failed to trigger module email: {e}")
        return False


def _build_file_image_html(fields: list, record_data: dict) -> str:
    """Build HTML section for file/image fields to append to email body."""
    parts = []

    for field in fields:
        field_type = field.get("type", "")
        field_name = field.get("name", "")
        field_label = field.get("label", field_name)
        value = record_data.get(field_name)

        if not value or field_type not in ("file", "image"):
            continue

        if field_type == "image":
            parts.append(
                f'<div style="margin:12px 0;">'
                f'<p style="font-weight:600;color:#374151;margin:0 0 8px 0;">{field_label}:</p>'
                f'<img src="{value}" alt="{field_label}" style="max-width:400px;border-radius:8px;border:1px solid #e5e7eb;" />'
                f'</div>'
            )
        elif field_type == "file":
            filename = value.split("/")[-1] if "/" in value else value
            parts.append(
                f'<div style="margin:12px 0;">'
                f'<p style="font-weight:600;color:#374151;margin:0 0 8px 0;">{field_label}:</p>'
                f'<a href="{value}" style="display:inline-block;padding:8px 16px;background:#3B82F6;color:#fff;'
                f'border-radius:6px;text-decoration:none;font-size:14px;">'
                f'📎 Download {filename}</a>'
                f'</div>'
            )

    if not parts:
        return ""

    return (
        '<div style="margin-top:24px;padding-top:16px;border-top:1px solid #e5e7eb;">'
        '<h3 style="font-size:16px;color:#1f2937;margin:0 0 12px 0;">📎 Attachments</h3>'
        + "".join(parts)
        + '</div>'
    )



# ============================================================
# Module Definition Endpoints
# ============================================================

@router.get("/clients/{client_id}/modules", response_model=PaginatedResponse[ClientModuleListResponse])
async def list_modules(
    client_id: int,
    search: str | None = Query(None, description="Filter by module name"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(PermissionChecker("module.view")),
    db: Session = Depends(get_db),
):
    """List dynamic modules for a client."""
    _get_client_or_404(client_id, db)

    query = db.query(ClientModule).filter(
        ClientModule.client_id == client_id,
        ClientModule.is_deleted == False,
    )
    if search:
        query = query.filter(ClientModule.name.ilike(f"%{search}%"))

    total = query.count()
    modules = query.order_by(ClientModule.created_at.desc()).offset(
        (page - 1) * page_size
    ).limit(page_size).all()

    # Batch count records — single query instead of N+1
    module_ids = [m.id for m in modules]
    record_counts: dict[int, int] = {}
    if module_ids:
        counts = db.query(
            ClientModuleRecord.module_id,
            func.count(ClientModuleRecord.id),
        ).filter(
            ClientModuleRecord.module_id.in_(module_ids),
            ClientModuleRecord.is_deleted == False,
        ).group_by(ClientModuleRecord.module_id).all()
        record_counts = dict(counts)

    items = [
        ClientModuleListResponse(
            id=m.id,
            client_id=m.client_id,
            name=m.name,
            slug=m.slug,
            description=m.description,
            icon=m.icon,
            field_count=len(m.fields) if m.fields else 0,
            record_count=record_counts.get(m.id, 0),
            fields=m.fields or [],
        )
        for m in modules
    ]
    return PaginatedResponse.create(items, total, page, page_size)


@router.get("/clients/{client_id}/modules/{module_id}", response_model=ClientModuleResponse)
async def get_module(
    client_id: int,
    module_id: int,
    current_user: User = Depends(PermissionChecker("module.view")),
    db: Session = Depends(get_db),
):
    """Get a specific module definition."""
    _get_client_or_404(client_id, db)
    module = db.query(ClientModule).filter(
        ClientModule.id == module_id,
        ClientModule.client_id == client_id,
        ClientModule.is_deleted == False,
    ).first()
    if not module:
        raise HTTPException(status_code=404, detail="Module not found")
    return _build_module_response(module, _get_record_count(module.id, db))


@router.post("/clients/{client_id}/modules", response_model=ClientModuleResponse, status_code=status.HTTP_201_CREATED)
async def create_module(
    client_id: int,
    data: ClientModuleCreate,
    current_user: User = Depends(PermissionChecker("module.create")),
    db: Session = Depends(get_db),
):
    """Create a new dynamic module with field definitions."""
    _get_client_or_404(client_id, db)

    slug = _slugify(data.name)
    existing = db.query(ClientModule).filter(
        ClientModule.client_id == client_id,
        ClientModule.slug == slug,
        ClientModule.is_deleted == False,
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Module '{data.name}' already exists for this client")

    module = ClientModule(
        client_id=client_id,
        name=data.name,
        slug=slug,
        description=data.description,
        icon=data.icon,
        fields=[f.model_dump() for f in data.fields],
        mail_template_id=data.mail_template_id,
        created_by=current_user.id,
    )
    db.add(module)
    db.commit()
    db.refresh(module)
    return _build_module_response(module, 0)


@router.post("/admin/quick-module", response_model=ClientModuleResponse, status_code=status.HTTP_201_CREATED)
async def quick_create_module(
    client_name: str,
    module_name: str,
    description: str | None = None,
    current_user: User = Depends(PermissionChecker("module.create")),
    db: Session = Depends(get_db),
):
    """
    Quickly create a module by providing Client Name and Module Name.
    Automatically finds the client and sets up default fields.
    """
    client = db.query(Client).filter(
        Client.name.ilike(f"%{client_name}%"),
        Client.is_deleted == False,
    ).first()
    if not client:
        client = db.query(Client).filter(
            Client.slug == _slugify(client_name),
            Client.is_deleted == False,
        ).first()
    if not client:
        raise HTTPException(status_code=404, detail=f"Client '{client_name}' not found")

    slug = _slugify(module_name)
    existing = db.query(ClientModule).filter(
        ClientModule.client_id == client.id,
        ClientModule.slug == slug,
        ClientModule.is_deleted == False,
    ).first()
    if existing:
        raise HTTPException(status_code=400, detail=f"Module '{module_name}' already exists for client '{client.name}'")

    default_fields = [
        {"name": "name", "label": "Name", "type": "text", "required": True, "placeholder": "Record name"},
        {"name": "status", "label": "Status", "type": "select", "options": ["New", "In Progress", "Done"], "required": True},
    ]

    module = ClientModule(
        client_id=client.id,
        name=module_name,
        slug=slug,
        description=description,
        icon="HiOutlineCollection",
        fields=default_fields,
        mail_template_id=None,
        created_by=current_user.id,
    )
    db.add(module)
    db.commit()
    db.refresh(module)
    return _build_module_response(module, 0)


@router.put("/clients/{client_id}/modules/{module_id}", response_model=ClientModuleResponse)
async def update_module(
    client_id: int,
    module_id: int,
    data: ClientModuleUpdate,
    current_user: User = Depends(PermissionChecker("module.edit")),
    db: Session = Depends(get_db),
):
    """Update a module definition."""
    _get_client_or_404(client_id, db)
    module = db.query(ClientModule).filter(
        ClientModule.id == module_id,
        ClientModule.client_id == client_id,
        ClientModule.is_deleted == False,
    ).first()
    if not module:
        raise HTTPException(status_code=404, detail="Module not found")

    update_data = data.model_dump(exclude_unset=True)

    if "name" in update_data:
        module.name = update_data["name"]
        module.slug = _slugify(update_data["name"])
    if "description" in update_data:
        module.description = update_data["description"]
    if "icon" in update_data:
        module.icon = update_data["icon"]
    if "fields" in update_data and data.fields is not None:
        module.fields = [f.model_dump() for f in data.fields]
    if "mail_template_id" in update_data:
        module.mail_template_id = update_data["mail_template_id"]

    module.updated_by = current_user.id
    db.commit()
    db.refresh(module)
    return _build_module_response(module, _get_record_count(module.id, db))


@router.delete("/clients/{client_id}/modules/{module_id}", response_model=MessageResponse)
async def delete_module(
    client_id: int,
    module_id: int,
    current_user: User = Depends(PermissionChecker("module.delete")),
    db: Session = Depends(get_db),
):
    """Delete a module and batch soft-delete all its records."""
    _get_client_or_404(client_id, db)
    module = db.query(ClientModule).filter(
        ClientModule.id == module_id,
        ClientModule.client_id == client_id,
        ClientModule.is_deleted == False,
    ).first()
    if not module:
        raise HTTPException(status_code=404, detail="Module not found")

    now = datetime.now(UTC)
    module.soft_delete(current_user.id)
    db.query(ClientModuleRecord).filter(
        ClientModuleRecord.module_id == module_id,
        ClientModuleRecord.is_deleted == False,
    ).update(
        {
            ClientModuleRecord.is_deleted: True,
            ClientModuleRecord.deleted_at: now,
            ClientModuleRecord.deleted_by: current_user.id,
        },
        synchronize_session="fetch",
    )
    db.commit()
    return MessageResponse(message="Module deleted successfully")


# ============================================================
# Module Record Endpoints
# ============================================================

@router.get("/clients/{client_id}/modules/{module_id}/records", response_model=PaginatedResponse[ClientModuleRecordResponse])
async def list_module_records(
    client_id: int,
    module_id: int,
    search: str | None = Query(None, description="Search within record data (JSON text match)"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(PermissionChecker("module.view")),
    db: Session = Depends(get_db),
):
    """List records for a module with optional text search."""
    _get_client_or_404(client_id, db)

    module = db.query(ClientModule).filter(
        ClientModule.id == module_id,
        ClientModule.client_id == client_id,
        ClientModule.is_deleted == False,
    ).first()
    if not module:
        raise HTTPException(status_code=404, detail="Module not found")

    columns = _get_table_columns(db, "client_module_records")
    required_columns = {"id", "module_id", "data"}
    if not required_columns.issubset(columns):
        logger.error("client_module_records missing required columns: %s", sorted(required_columns - columns))
        raise HTTPException(status_code=500, detail="Client module records table is not ready")

    select_columns = ["id", "module_id", "data"]
    select_columns.append("created_at" if "created_at" in columns else "NULL AS created_at")
    select_columns.append("updated_at" if "updated_at" in columns else "NULL AS updated_at")

    where_parts = ["module_id = :module_id"]
    params = {
        "module_id": module_id,
        "limit": page_size,
        "offset": (page - 1) * page_size,
    }

    if "is_deleted" in columns:
        where_parts.append("(is_deleted = 0 OR is_deleted IS NULL)")
    if search:
        # JSON text search — works for MySQL JSON columns via LIKE on cast
        where_parts.append("CAST(data AS CHAR) LIKE :search")
        params["search"] = f"%{search}%"

    where_clause = " AND ".join(where_parts)
    order_clause = "created_at DESC, id DESC" if "created_at" in columns else "id DESC"

    total = db.execute(
        text(f"SELECT COUNT(*) FROM client_module_records WHERE {where_clause}"),  # noqa: S608
        params,
    ).scalar() or 0
    records_sql = (  # noqa: S608
        "SELECT "  # noqa: S608
        + ", ".join(select_columns)
        + f" FROM client_module_records WHERE {where_clause}"
        + f" ORDER BY {order_clause} LIMIT :limit OFFSET :offset"
    )
    records = db.execute(text(records_sql), params).all()  # noqa: S608

    return PaginatedResponse.create(
        [_legacy_safe_record_response(r) for r in records],
        total, page, page_size,
    )


@router.get("/clients/{client_id}/modules/{module_id}/records/export")
async def export_module_records(
    client_id: int,
    module_id: int,
    format: str = Query("csv", description="Export format: csv, excel, pdf"),
    current_user: User = Depends(PermissionChecker("module.view")),
    db: Session = Depends(get_db),
):
    """Export all records for a module as CSV, Excel, or PDF."""
    _get_client_or_404(client_id, db)

    module = db.query(ClientModule).filter(
        ClientModule.id == module_id,
        ClientModule.client_id == client_id,
        ClientModule.is_deleted == False,
    ).first()
    if not module:
        raise HTTPException(status_code=404, detail="Module not found")

    records = db.query(ClientModuleRecord).filter(
        ClientModuleRecord.module_id == module_id,
        ClientModuleRecord.is_deleted == False,
    ).order_by(ClientModuleRecord.created_at.desc()).all()

    fields = module.fields or []
    field_names = [f["name"] for f in fields]
    field_labels = [f.get("label", f["name"]) for f in fields]

    # Build rows
    rows = []
    for r in records:
        row = [str(r.data.get(fn, "")) if r.data else "" for fn in field_names]
        row.append(r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "")
        rows.append(row)
    headers = field_labels + ["Created At"]

    safe_name = re.sub(r'[^a-zA-Z0-9_-]', '_', module.name)

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}_records.csv"'},
        )

    elif format == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active
        ws.title = module.name[:31]  # Excel sheet name max 31 chars

        # Style header row
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side(style="thin", color="E5E7EB"),
            right=Side(style="thin", color="E5E7EB"),
            top=Side(style="thin", color="E5E7EB"),
            bottom=Side(style="thin", color="E5E7EB"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}_records.xlsx"'},
        )

    elif format == "pdf":
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=15*mm, bottomMargin=15*mm)
        styles = getSampleStyleSheet()
        elements = []

        # Title
        title_style = styles["Title"]
        title_style.fontSize = 16
        elements.append(Paragraph(f"{module.name} — Records", title_style))
        elements.append(Spacer(1, 10*mm))

        # Table data
        table_data = [headers]
        for row in rows:
            # Wrap long text for PDF
            table_data.append([
                Paragraph(str(v)[:80], styles["Normal"]) if len(str(v)) > 30 else str(v)
                for v in row
            ])

        num_cols = len(headers)
        available_width = landscape(A4)[0] - 30*mm
        col_width = available_width / num_cols

        t = Table(table_data, colWidths=[col_width] * num_cols, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ]))
        elements.append(t)

        doc.build(elements)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{safe_name}_records.pdf"'},
        )

    else:
        raise HTTPException(status_code=400, detail="Invalid format. Use csv, excel, or pdf.")


@router.post("/clients/{client_id}/modules/{module_id}/records", response_model=ClientModuleRecordResponse, status_code=status.HTTP_201_CREATED)
async def create_module_record(
    client_id: int,
    module_id: int,
    data: ClientModuleRecordCreate,
    current_user: User = Depends(PermissionChecker("module.create")),
    db: Session = Depends(get_db),
):
    """Create a record in a module."""
    _get_client_or_404(client_id, db)
    module = db.query(ClientModule).filter(
        ClientModule.id == module_id,
        ClientModule.client_id == client_id,
        ClientModule.is_deleted == False,
    ).first()
    if not module:
        raise HTTPException(status_code=404, detail="Module not found")

    _validate_required_fields(module, data.data)

    record = ClientModuleRecord(
        module_id=module_id,
        data=data.data,
        created_by=current_user.id,
    )
    db.add(record)
    db.commit()
    db.refresh(record)
    email_sent = await _trigger_module_email(client_id, module, record.data, db)
    return _build_record_response(record, email_sent)


@router.put("/clients/{client_id}/modules/{module_id}/records/{record_id}", response_model=ClientModuleRecordResponse)
async def update_module_record(
    client_id: int,
    module_id: int,
    record_id: int,
    data: ClientModuleRecordUpdate,
    current_user: User = Depends(PermissionChecker("module.edit")),
    db: Session = Depends(get_db),
):
    """Update a module record."""
    _get_client_or_404(client_id, db)
    record = db.query(ClientModuleRecord).filter(
        ClientModuleRecord.id == record_id,
        ClientModuleRecord.module_id == module_id,
        ClientModuleRecord.is_deleted == False,
    ).first()
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")

    record.data = data.data
    record.updated_by = current_user.id
    db.commit()
    db.refresh(record)
    return _build_record_response(record)


@router.delete("/clients/{client_id}/modules/{module_id}/records/{record_id}", response_model=MessageResponse)
async def delete_module_record(
    client_id: int,
    module_id: int,
    record_id: int,
    current_user: User = Depends(PermissionChecker("module.delete")),
    db: Session = Depends(get_db),
):
    """Delete a module record."""
    _get_client_or_404(client_id, db)
    record = db.query(ClientModuleRecord).filter(
        ClientModuleRecord.id == record_id,
        ClientModuleRecord.module_id == module_id,
        ClientModuleRecord.is_deleted == False,
    ).first()
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")

    record.soft_delete(current_user.id)
    db.commit()
    return MessageResponse(message="Record deleted successfully")


# ============================================================
# API Key Management
# ============================================================

@router.post("/clients/{client_id}/api-key", response_model=dict)
async def generate_api_key(
    client_id: int,
    current_user: User = Depends(PermissionChecker("module.edit")),
    db: Session = Depends(get_db),
):
    """Generate or regenerate an API key for a client."""
    client = _get_client_or_404(client_id, db)
    client.api_key = secrets.token_hex(32)
    client.updated_by = current_user.id
    db.commit()
    db.refresh(client)
    return {
        "api_key": client.api_key,
        "message": "API key generated successfully. Store it safely — it won't be shown again.",
    }


@router.get("/clients/{client_id}/api-key", response_model=dict)
async def get_api_key_status(
    client_id: int,
    current_user: User = Depends(PermissionChecker("module.view")),
    db: Session = Depends(get_db),
):
    """Check if an API key exists (does not reveal the key)."""
    client = _get_client_or_404(client_id, db)
    preview = (
        f"{client.api_key[:8]}...{client.api_key[-4:]}"
        if client.api_key else None
    )
    return {"has_api_key": bool(client.api_key), "api_key_preview": preview}


@router.delete("/clients/{client_id}/api-key", response_model=MessageResponse)
async def revoke_api_key(
    client_id: int,
    current_user: User = Depends(PermissionChecker("module.edit")),
    db: Session = Depends(get_db),
):
    """Revoke the API key for a client."""
    client = _get_client_or_404(client_id, db)
    client.api_key = None
    client.updated_by = current_user.id
    db.commit()
    return MessageResponse(message="API key revoked successfully")


# ============================================================
# Public API Endpoints (X-API-Key header auth)
# ============================================================

def _auth_by_api_key(api_key: str, db: Session) -> Client:
    """Validate bare API key and return client."""
    if not api_key:
        raise HTTPException(status_code=401, detail="API key is required")
    client = db.query(Client).filter(
        Client.api_key == api_key,
        Client.is_deleted == False,
    ).first()
    if not client:
        raise HTTPException(status_code=401, detail="Invalid API key")
    return client


def _auth_by_slug_and_key(client_slug: str, api_key: str, db: Session) -> Client:
    """Validate client slug + API key pair."""
    if not api_key:
        raise HTTPException(status_code=401, detail="API key is required")
    # Sanitise slug input
    if not re.match(r'^[\w-]+$', client_slug):
        raise HTTPException(status_code=400, detail="Invalid client slug")
    client = db.query(Client).filter(
        Client.slug == client_slug,
        Client.is_deleted == False,
    ).first()
    if not client or client.api_key != api_key:
        raise HTTPException(status_code=401, detail="Invalid client slug or API key")
    return client


def _get_public_module(client: Client, module_slug: str, db: Session) -> ClientModule:
    """Fetch a public module by slug, with input sanitisation."""
    if not re.match(r'^[\w-]+$', module_slug):
        raise HTTPException(status_code=400, detail="Invalid module slug")
    module = db.query(ClientModule).filter(
        ClientModule.client_id == client.id,
        ClientModule.slug == module_slug,
        ClientModule.is_deleted == False,
    ).first()
    if not module:
        raise HTTPException(status_code=404, detail="Module not found")
    return module


# ---- Simplified Routes (API key identifies client) ----

@router.get("/public/modules")
async def public_list_modules_simple(
    x_api_key: str = Header(..., alias="X-API-Key"),
    db: Session = Depends(get_db),
):
    """List all modules for the authenticated client (simple URL)."""
    client = _auth_by_api_key(x_api_key, db)
    return await _public_modules_response(client, db)


@router.get("/public/{module_slug}/records")
async def public_list_records_simple(
    module_slug: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    x_api_key: str = Header(..., alias="X-API-Key"),
    db: Session = Depends(get_db),
):
    """List records for a module (simple URL)."""
    client = _auth_by_api_key(x_api_key, db)
    module = _get_public_module(client, module_slug, db)
    return await _public_records_response(client, module, page, page_size, db)


@router.post("/public/{module_slug}/records", status_code=status.HTTP_201_CREATED)
async def public_create_record_simple(
    module_slug: str,
    data: ClientModuleRecordCreate,
    x_api_key: str = Header(..., alias="X-API-Key"),
    db: Session = Depends(get_db),
):
    """Create a record (simple URL)."""
    client = _auth_by_api_key(x_api_key, db)
    module = _get_public_module(client, module_slug, db)
    return await _public_create_record(client, module, data, db)


# ---- Slug Routes (client_slug in URL for multi-tenant clarity) ----

@router.get("/public/{client_slug}/modules")
async def public_list_modules_slug(
    client_slug: str,
    x_api_key: str = Header(..., alias="X-API-Key"),
    db: Session = Depends(get_db),
):
    """List all modules for a specific client (slug URL)."""
    client = _auth_by_slug_and_key(client_slug, x_api_key, db)
    return await _public_modules_response(client, db)


@router.get("/public/{client_slug}/{module_slug}/records")
async def public_list_records_slug(
    client_slug: str,
    module_slug: str,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    x_api_key: str = Header(..., alias="X-API-Key"),
    db: Session = Depends(get_db),
):
    """List records (slug URL)."""
    client = _auth_by_slug_and_key(client_slug, x_api_key, db)
    module = _get_public_module(client, module_slug, db)
    return await _public_records_response(client, module, page, page_size, db)


@router.post("/public/{client_slug}/{module_slug}/records", status_code=status.HTTP_201_CREATED)
async def public_create_record_slug(
    client_slug: str,
    module_slug: str,
    data: ClientModuleRecordCreate,
    x_api_key: str = Header(..., alias="X-API-Key"),
    db: Session = Depends(get_db),
):
    """Create a record (slug URL)."""
    client = _auth_by_slug_and_key(client_slug, x_api_key, db)
    module = _get_public_module(client, module_slug, db)
    return await _public_create_record(client, module, data, db)


# ---- Public Shared Implementations ----

async def _public_modules_response(client: Client, db: Session) -> dict:
    modules = db.query(ClientModule).filter(
        ClientModule.client_id == client.id,
        ClientModule.is_deleted == False,
    ).order_by(ClientModule.created_at.desc()).all()
    return {
        "client": {"id": client.id, "name": client.name, "slug": client.slug},
        "modules": [
            {
                "id": m.id,
                "name": m.name,
                "slug": m.slug,
                "description": m.description,
                "fields": m.fields or [],
            }
            for m in modules
        ],
    }


async def _public_records_response(client: Client, module: ClientModule, page: int, page_size: int, db: Session) -> dict:
    query = db.query(ClientModuleRecord).filter(
        ClientModuleRecord.module_id == module.id,
        ClientModuleRecord.is_deleted == False,
    )
    total = query.count()
    records = query.order_by(ClientModuleRecord.created_at.desc()).offset(
        (page - 1) * page_size
    ).limit(page_size).all()
    return {
        "client": {"id": client.id, "name": client.name, "slug": client.slug},
        "module": {"id": module.id, "name": module.name, "slug": module.slug},
        "total": total,
        "page": page,
        "page_size": page_size,
        "records": [
            {
                "id": r.id,
                "data": r.data or {},
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
            for r in records
        ],
    }


async def _public_create_record(client: Client, module: ClientModule, data: ClientModuleRecordCreate, db: Session) -> dict:
    _validate_required_fields(module, data.data)
    record = ClientModuleRecord(module_id=module.id, data=data.data)
    db.add(record)
    db.commit()
    db.refresh(record)
    email_sent = await _trigger_module_email(client.id, module, record.data, db)
    return {
        "id": record.id,
        "module": module.name,
        "data": record.data,
        "email_sent": email_sent,
        "created_at": record.created_at.isoformat() if record.created_at else None,
        "message": "Record created successfully",
    }
