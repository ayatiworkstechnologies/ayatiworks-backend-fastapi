"""
Invoice and Billing API routes.
"""

from datetime import date
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, Query, status
from sqlalchemy import func
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.api.deps import PermissionChecker
from app.core.exceptions import ResourceNotFoundError
from app.database import get_db
from app.models.auth import User
from app.models.invoice import Invoice, InvoiceItem, InvoiceStatus, Payment
from app.schemas.common import MessageResponse, PaginatedResponse
from app.schemas.invoice import (
    InvoiceCreate,
    InvoiceListResponse,
    InvoiceResponse,
    InvoiceUpdate,
    PaymentCreate,
    PaymentResponse,
)

router = APIRouter(prefix="/invoices", tags=["Invoices"])


def generate_invoice_number(db: Session) -> str:
    """Generate next invoice number."""
    year = date.today().year
    prefix = f"INV-{year}-"

    last = db.query(func.max(Invoice.invoice_number)).filter(
        Invoice.invoice_number.like(f"{prefix}%")
    ).scalar()

    if last:
        try:
            num = int(last.replace(prefix, "")) + 1
        except ValueError:
            num = 1
    else:
        num = 1

    return f"{prefix}{num:04d}"


@router.get("", response_model=PaginatedResponse[InvoiceListResponse])
async def list_invoices(
    client_id: int | None = None,
    company_id: int | None = None,
    status: str | None = None,
    search: str | None = None,
    from_date: date | None = None,
    to_date: date | None = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(PermissionChecker("invoice.view")),
    db: Session = Depends(get_db)
):
    """List all invoices with filters, search, and pagination."""
    from app.models.client import Client

    query = db.query(Invoice).filter(Invoice.is_deleted == False)

    if company_id:
        query = query.filter(Invoice.company_id == company_id)

    if client_id:
        query = query.filter(Invoice.client_id == client_id)

    if status:
        query = query.filter(Invoice.status == status)

    if from_date:
        query = query.filter(Invoice.issue_date >= from_date)

    if to_date:
        query = query.filter(Invoice.issue_date <= to_date)

    if search:
        search_term = f"%{search}%"
        query = query.outerjoin(Client, Invoice.client_id == Client.id).filter(
            Invoice.invoice_number.ilike(search_term)
            | Client.name.ilike(search_term)
        )

    total = query.count()

    offset = (page - 1) * page_size
    invoices = query.order_by(Invoice.issue_date.desc()).offset(offset).limit(page_size).all()

    items = []
    for inv in invoices:
        items.append(InvoiceListResponse(
            id=inv.id,
            invoice_number=inv.invoice_number,
            client_name=inv.client.name if inv.client else "",
            issue_date=inv.issue_date,
            due_date=inv.due_date,
            total=inv.total,
            amount_due=inv.amount_due,
            status=inv.status
        ))

    return PaginatedResponse.create(items, total, page, page_size)


@router.get("/{invoice_id}", response_model=InvoiceResponse)
async def get_invoice(
    invoice_id: int,
    current_user: User = Depends(PermissionChecker("invoice.view")),
    db: Session = Depends(get_db)
):
    """Get invoice by ID."""
    invoice = db.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.is_deleted == False
    ).first()

    if not invoice:
        raise ResourceNotFoundError("Invoice", invoice_id)

    response = InvoiceResponse.model_validate(invoice)
    response.client_name = invoice.client.name if invoice.client else None

    return response


@router.post("", response_model=InvoiceResponse, status_code=status.HTTP_201_CREATED)
async def create_invoice(
    data: InvoiceCreate,
    current_user: User = Depends(PermissionChecker("invoice.create")),
    db: Session = Depends(get_db)
):
    """Create a new invoice."""
    # Calculate totals
    subtotal = Decimal(0)
    for item in data.items:
        subtotal += item.rate * Decimal(str(item.quantity))

    discount = data.discount
    if data.discount_type == "percentage":
        discount = subtotal * (data.discount / 100)

    tax = (subtotal - discount) * Decimal(str(data.tax_rate / 100))
    total = subtotal - discount + tax

    for attempt in range(5):
        invoice_number = generate_invoice_number(db)

        invoice = Invoice(
            invoice_number=invoice_number,
            client_id=data.client_id,
            project_id=data.project_id,
            reference=data.reference,
            issue_date=data.issue_date,
            due_date=data.due_date,
            subtotal=subtotal,
            discount=discount,
            discount_type=data.discount_type,
            tax=tax,
            tax_rate=data.tax_rate,
            total=total,
            amount_due=total,
            currency=data.currency,
            notes=data.notes,
            terms=data.terms,
            created_by=current_user.id
        )

        db.add(invoice)
        try:
            db.flush()
        except IntegrityError as exc:
            db.rollback()
            if "invoice_number" not in str(exc).lower() or attempt == 4:
                raise
            continue

        # Add invoice items
        for i, item_data in enumerate(data.items):
            item = InvoiceItem(
                invoice_id=invoice.id,
                description=item_data.description,
                quantity=item_data.quantity,
                rate=item_data.rate,
                amount=item_data.rate * Decimal(str(item_data.quantity)),
                hours=item_data.hours,
                order=i
            )
            db.add(item)

        db.commit()
        db.refresh(invoice)

        return InvoiceResponse.model_validate(invoice)

    raise HTTPException(
        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        detail="Failed to generate a unique invoice number",
    )


@router.put("/{invoice_id}", response_model=InvoiceResponse)
async def update_invoice(
    invoice_id: int,
    data: InvoiceUpdate,
    current_user: User = Depends(PermissionChecker("invoice.edit")),
    db: Session = Depends(get_db)
):
    """Update an invoice."""
    invoice = db.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.is_deleted == False
    ).first()

    if not invoice:
        raise ResourceNotFoundError("Invoice", invoice_id)

    # Don't allow editing paid invoices
    if invoice.status == InvoiceStatus.PAID.value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot edit a paid invoice"
        )

    # Update fields
    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        if hasattr(invoice, field) and field not in ['items']:
            setattr(invoice, field, value)

    if "discount" in update_data or "tax_rate" in update_data:
        taxable = invoice.subtotal - invoice.discount
        if taxable < 0:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Discount cannot exceed subtotal",
            )
        invoice.tax = taxable * Decimal(str(invoice.tax_rate / 100))
        invoice.total = taxable + invoice.tax
        invoice.amount_due = invoice.total - invoice.amount_paid

    db.commit()
    db.refresh(invoice)

    return InvoiceResponse.model_validate(invoice)


@router.delete("/{invoice_id}", response_model=MessageResponse)
async def delete_invoice(
    invoice_id: int,
    current_user: User = Depends(PermissionChecker("invoice.delete")),
    db: Session = Depends(get_db)
):
    """Delete an invoice (soft delete)."""
    invoice = db.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.is_deleted == False
    ).first()

    if not invoice:
        raise ResourceNotFoundError("Invoice", invoice_id)

    # Don't allow deleting paid invoices
    if invoice.status == InvoiceStatus.PAID.value:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot delete a paid invoice"
        )

    invoice.is_deleted = True
    db.commit()

    return MessageResponse(message="Invoice deleted successfully")


@router.post("/{invoice_id}/send", response_model=MessageResponse)
async def send_invoice(
    invoice_id: int,
    current_user: User = Depends(PermissionChecker("invoice.edit")),
    db: Session = Depends(get_db)
):
    """Mark invoice as sent."""
    invoice = db.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.is_deleted == False
    ).first()

    if not invoice:
        raise ResourceNotFoundError("Invoice", invoice_id)

    from datetime import datetime
    invoice.status = InvoiceStatus.SENT.value
    invoice.sent_at = datetime.utcnow()

    db.commit()

    # TODO: Send email to client

    return MessageResponse(message="Invoice sent successfully")


@router.post("/{invoice_id}/payments", response_model=PaymentResponse, status_code=status.HTTP_201_CREATED)
async def record_payment(
    invoice_id: int,
    data: PaymentCreate,
    current_user: User = Depends(PermissionChecker("invoice.edit")),
    db: Session = Depends(get_db)
):
    """Record a payment for an invoice."""
    invoice = db.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.is_deleted == False
    ).first()

    if not invoice:
        raise ResourceNotFoundError("Invoice", invoice_id)

    if data.amount <= 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Payment amount must be positive",
        )
    if invoice.status == InvoiceStatus.PAID.value or invoice.amount_due <= 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invoice is already fully paid",
        )
    if data.amount > invoice.amount_due:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Payment exceeds amount due",
        )

    payment = Payment(
        invoice_id=invoice_id,
        amount=data.amount,
        currency=invoice.currency,
        payment_date=data.payment_date,
        payment_method=data.payment_method,
        reference=data.reference,
        notes=data.notes,
        created_by=current_user.id
    )

    db.add(payment)

    # Update invoice
    invoice.amount_paid += data.amount
    invoice.amount_due -= data.amount

    if invoice.amount_due <= 0:
        invoice.status = InvoiceStatus.PAID.value
    else:
        invoice.status = InvoiceStatus.PARTIAL.value

    db.commit()
    db.refresh(payment)

    return PaymentResponse.model_validate(payment)


@router.get("/{invoice_id}/payments", response_model=list[PaymentResponse])
async def get_invoice_payments(
    invoice_id: int,
    current_user: User = Depends(PermissionChecker("invoice.view")),
    db: Session = Depends(get_db)
):
    """Get all payments for an invoice."""
    payments = db.query(Payment).filter(
        Payment.invoice_id == invoice_id,
        Payment.is_deleted == False
    ).order_by(Payment.payment_date).all()

    return [PaymentResponse.model_validate(p) for p in payments]


@router.get("/export/{format}")
async def export_invoices(
    format: str,
    client_id: int | None = None,
    status: str | None = None,
    from_date: date | None = None,
    to_date: date | None = None,
    current_user: User = Depends(PermissionChecker("invoice.view")),
    db: Session = Depends(get_db)
):
    """
    Export invoices to CSV, Excel, or PDF.
    Requires invoice.view permission.
    """
    import csv
    import io

    from fastapi.responses import StreamingResponse

    if format not in ("csv", "excel", "pdf"):
        raise HTTPException(status_code=400, detail="Format must be csv, excel, or pdf")

    query = db.query(Invoice).filter(Invoice.is_deleted == False)
    if client_id:
        query = query.filter(Invoice.client_id == client_id)
    if status:
        query = query.filter(Invoice.status == status)
    if from_date:
        query = query.filter(Invoice.issue_date >= from_date)
    if to_date:
        query = query.filter(Invoice.issue_date <= to_date)

    invoices = query.order_by(Invoice.issue_date.desc()).limit(10000).all()

    headers = ["Invoice #", "Client", "Issue Date", "Due Date", "Subtotal", "Tax", "Total", "Amount Due", "Status"]
    rows = []
    for inv in invoices:
        rows.append([
            inv.invoice_number,
            inv.client.name if inv.client else "",
            str(inv.issue_date) if inv.issue_date else "",
            str(inv.due_date) if inv.due_date else "",
            f"{inv.subtotal:.2f}" if inv.subtotal else "0.00",
            f"{inv.tax:.2f}" if inv.tax else "0.00",
            f"{inv.total:.2f}" if inv.total else "0.00",
            f"{inv.amount_due:.2f}" if inv.amount_due else "0.00",
            inv.status or "",
        ])

    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(rows)
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=invoices.csv"}
        )

    if format == "excel":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Invoices"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for r, row in enumerate(rows, 2):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=invoices.xlsx"}
        )

    # PDF
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    table_data = [headers] + rows
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ("ALIGN", (4, 1), (-2, -1), "RIGHT"),
    ]))
    doc.build([table])
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=invoices.pdf"}
    )


